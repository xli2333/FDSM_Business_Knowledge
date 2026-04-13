"""
生成KPI考核表Excel文件
格式参照截图：业务指标(70%) + 工作态度指标(20%) + 个人能力提升指标(10%)
"""
import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "KPI考核表"

# ── 样式定义 ──
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
hdr_font = Font(name='微软雅黑', size=14, bold=True)
sec_font = Font(name='微软雅黑', size=11, bold=True, color='C00000')
lbl_font = Font(name='微软雅黑', size=10, bold=True)
txt_font = Font(name='微软雅黑', size=10)
sm_font  = Font(name='微软雅黑', size=9)
col_hdr_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
blue_fill  = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
gray_fill  = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
dark_fill  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
wrap  = Alignment(wrap_text=True, vertical='top', horizontal='left')
ctr   = Alignment(wrap_text=True, vertical='center', horizontal='center')
lctr  = Alignment(wrap_text=True, vertical='center', horizontal='left')

# 列宽
for col, w in {'A': 5, 'B': 50, 'C': 50, 'D': 10}.items():
    ws.column_dimensions[col].width = w


def write_section_header(r, text):
    """大分类标题行，如：业务指标（权重 70%）"""
    ws.merge_cells(f'A{r}:D{r}')
    ws[f'A{r}'].value = text
    ws[f'A{r}'].font = sec_font
    ws[f'A{r}'].alignment = lctr
    for c in 'ABCD':
        ws[f'{c}{r}'].fill = blue_fill
        ws[f'{c}{r}'].border = thin
    ws.row_dimensions[r].height = 28
    return r + 1


def write_sub_objective(r, label, obj_text, weight_text):
    """子目标行，如：*1. 业务指标  描述...  权重 65%"""
    ws[f'A{r}'].value = label
    ws[f'A{r}'].font = lbl_font
    ws[f'A{r}'].alignment = ctr
    ws[f'A{r}'].border = thin
    ws.merge_cells(f'B{r}:C{r}')
    ws[f'B{r}'].value = obj_text
    ws[f'B{r}'].font = txt_font
    ws[f'B{r}'].alignment = wrap
    ws[f'B{r}'].border = thin
    ws[f'C{r}'].border = thin
    ws[f'D{r}'].value = weight_text
    ws[f'D{r}'].font = lbl_font
    ws[f'D{r}'].alignment = ctr
    ws[f'D{r}'].border = thin
    ws.row_dimensions[r].height = 48
    return r + 1


def write_col_headers(r):
    for i, h in enumerate(['序号', '关键举措', '衡量标准', '权重(%)']):
        cell = ws[f'{get_column_letter(i+1)}{r}']
        cell.value = h
        cell.font = col_hdr_font
        cell.fill = dark_fill
        cell.alignment = ctr
        cell.border = thin
    ws.row_dimensions[r].height = 28
    return r + 1


def write_items(r, items):
    for idx, item in enumerate(items, 1):
        ws[f'A{r}'].value = idx
        ws[f'A{r}'].font = txt_font
        ws[f'A{r}'].alignment = ctr
        ws[f'A{r}'].border = thin
        ws[f'B{r}'].value = item['desc']
        ws[f'B{r}'].font = txt_font
        ws[f'B{r}'].alignment = wrap
        ws[f'B{r}'].border = thin
        ws[f'C{r}'].value = item['metrics']
        ws[f'C{r}'].font = txt_font
        ws[f'C{r}'].alignment = wrap
        ws[f'C{r}'].border = thin
        ws[f'D{r}'].value = item['weight']
        ws[f'D{r}'].font = txt_font
        ws[f'D{r}'].alignment = ctr
        ws[f'D{r}'].border = thin
        lines = max(item['desc'].count('\n'), item['metrics'].count('\n')) + 3
        ws.row_dimensions[r].height = max(lines * 16, 100)
        if idx % 2 == 0:
            for c in 'ABCD':
                ws[f'{c}{r}'].fill = gray_fill
        r += 1
    return r


# ══════════════════════════════════════════════
# 标题 + 基本信息
# ══════════════════════════════════════════════
ws.merge_cells('A1:D1')
ws['A1'].value = "KPI考核表"
ws['A1'].font = hdr_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

ws['A2'].value = '考核期间'
ws['A2'].font = lbl_font
ws['A2'].alignment = ctr
ws['B2'].value = '2026-01-01 ~ 2027-01-31'
ws['B2'].font = txt_font
ws['B2'].alignment = lctr
ws['C2'].value = '部门：KDCC'
ws['C2'].font = txt_font
ws['C2'].alignment = lctr
ws.row_dimensions[2].height = 24
ws.row_dimensions[3].height = 10

# ══════════════════════════════════════════════════
# 第一部分：业务指标（权重 70%）
# ══════════════════════════════════════════════════
row = write_section_header(4, '业务指标（权重 70%）    70%')

# ── *1. 业务指标：知识平台  权重 65% ──
row = write_sub_objective(row,
    '*1. 业务指标',
    '复旦管院知识库：完成知识平台的搭建与上线、内容整理与运营、与IT部门的协作对接、以及初期用户验证等核心产出目标',
    '权重 65%')

row = write_col_headers(row)

biz1_items = [
    {
        'desc': (
            '知识平台搭建与上线\n\n'
            '完成"复旦管院智识库"网站的搭建与正式上线，主要包括：\n'
            '• 智能搜索功能：用户输入关键词或一句话，即可精准找到相关文章\n'
            '• AI问答助手\n'
            '• 内容分类浏览\n'
            '• 会员体系构建\n'
            '• 以及其他可能需要的功能'
        ),
        'metrics': (
            '1. 网站主要功能开发完成并可正常使用\n'
            '2. 搜索准确度高：用户搜索后，前3条结果中有想要内容的概率 > 85%\n'
            '3. 管院知识资产导入平台'
        ),
        'weight': 15,
    },
    {
        'desc': (
            '与IT部门协作对接\n\n'
            '与学校IT部门紧密配合，完成平台上线所需的各项对接工作：\n'
            '• 网站域名申请与上线（让用户能通过网址访问）\n'
            '• 服务器环境部署配合（IT部门提供运行环境支持）\n'
            '• 日常运维保障对接（建立故障响应机制）'
        ),
        'metrics': (
            '1. 与IT部门定期会议按时召开率 ≥ 90%\n'
            '2. 域名上线等关键节点按时完成\n'
            '3. 信息安全审查全部通过\n'
            '4. 运维交接文档完整交付'
        ),
        'weight': 15,
    },
    {
        'desc': (
            '内容管理后台搭建\n\n'
            '为编辑团队开发一个方便操作的内容管理后台：\n'
            '• 编辑可在后台创建、修改、审核、发布文章，全流程线上完成\n'
            '• AI辅助功能：自动为文章生成标签、摘要，减少人工操作\n'
            '• 可为文章一键智能排版\n'
            '• 管理员权限分级，操作留痕可追溯'
        ),
        'metrics': (
            '1. 后台核心功能交付完成\n'
            '2. 编辑发布一篇文章的平均耗时 < 15分钟'
        ),
        'weight': 10,
    },
    {
        'desc': (
            '内容运营体系搭建\n\n'
            '建立智识库内容运营的标准化流程：\n'
            '• 制定不同栏目的内容策划方案\n'
            '• 规划不同核心主题的内容方向（AI、ESG、领导力、数字化等）\n'
            '• 制定每周内容更新计划并执行\n'
            '• 建立内容质量评估标准'
        ),
        'metrics': (
            '1. 存量文章分类标签整理完成率 > 95%\n'
            '2. 各个主题栏目均有内容覆盖，无明显空白'
        ),
        'weight': 10,
    },
    {
        'desc': (
            '用户测试与初期验证（起步阶段）\n\n'
            '重点是验证产品是否好用、用户是否需要：\n'
            '• 根据试用用户反馈持续优化产品体验\n'
            '• 邀请校友种子用户，验证付费意愿\n'
            '• 搭建用户行为数据统计，了解哪些功能最受欢迎'
        ),
        'metrics': (
            '1. 首批测试用户数量 ≥ 50人\n'
            '2. 主要功能的使用率 > 60%（即大多数用户会用到核心功能）\n'
            '3. 完成 ≥ 3轮用户反馈迭代'
        ),
        'weight': 15,
    },
]

row = write_items(row, biz1_items)

# ── *2. 业务指标：失败案例  权重 5% ──
row += 1
row = write_sub_objective(row,
    '*2. 业务指标',
    '失败案例创作与失败知识图谱网站构建',
    '权重 5%')

row = write_col_headers(row)

biz2_items = [
    {
        'desc': (
            '失败案例收集与失败知识图谱构建\n\n'
            '系统性收集和整理商业失败案例，构建"失败知识图谱"：\n'
            '• 收集整理国内外典型商业失败案例，涵盖创业失败、战略失误、管理危机等\n'
            '• 构建失败案例之间的关联关系图谱，用户可按地区、行业、产品等维度检索浏览'
        ),
        'metrics': (
            '完成失败案例与教学参考的编写，完成知识图谱网站的构建'
        ),
        'weight': 5,
    },
]

row = write_items(row, biz2_items)

# ══════════════════════════════════════════════════
# 第二部分：工作态度指标（权重 20%）
# ══════════════════════════════════════════════════
row += 1
row = write_section_header(row, '工作态度指标（权重 20%）    20%')

row = write_sub_objective(row,
    '*1. 工作态度指标',
    '在智识库项目推进、跨部门协作与项目交付中保持高度责任心、主动性与协作精神，确保工作质量与时效性',
    '权重 20%')

row = write_col_headers(row)

attitude_items = [
    {
        'desc': (
            '项目交付时效与责任心\n\n'
            '严格遵守项目节点与交付时间要求，对所负责的智识库平台搭建、'
            'IT部门对接、内容运营、用户测试等各项工作的进度与质量承担主体责任'
        ),
        'metrics': (
            '1. 各项目关键节点按时完成率 ≥ 95%\n'
            '2. 因个人原因导致的延期次数 ≤ 1次/季度\n'
            '3. 交付成果一次通过率（无需反复返工）'
        ),
        'weight': 6,
    },
    {
        'desc': (
            '跨部门协作与沟通\n\n'
            '积极配合IT部门、编辑团队等多方协作需求，在平台开发对接、'
            '内容运营配合、用户测试组织等工作中保持高效沟通与主动响应'
        ),
        'metrics': (
            '1. 协作方满意度反馈（IT部门/编辑团队/内部团队）\n'
            '2. 跨部门需求响应及时性（24小时内响应）\n'
            '3. 协作过程中的主动沟通与问题预警表现'
        ),
        'weight': 5,
    },
    {
        'desc': (
            '主动担当与问题解决\n\n'
            '面对平台开发中的技术难题、IT对接中的协调困难、'
            '内容运营中的挑战性任务，主动探索解决方案，遇到困难不回避、不推诿，积极推动工作向前'
        ),
        'metrics': (
            '1. 主动提出改进建议或创新方案的次数\n'
            '2. 遇到阻碍时的应对方式与推进速度\n'
            '3. 上级/同事对工作主动性的评价'
        ),
        'weight': 5,
    },
    {
        'desc': (
            '质量意识与细节把控\n\n'
            '在平台功能开发、内容整理、用户体验优化、文档编写等工作中始终坚持高质量标准，注重细节与规范性'
        ),
        'metrics': (
            '1. 成果质量评审得分达标率\n'
            '2. 因质量问题被退回或返工的比例 ≤ 5%\n'
            '3. 内部用户/合作方对交付质量的满意度'
        ),
        'weight': 4,
    },
]

row = write_items(row, attitude_items)

# ══════════════════════════════════════════════════
# 第三部分：个人能力提升指标（权重 10%）
# ══════════════════════════════════════════════════
row += 1
row = write_section_header(row, '个人能力提升指标（权重 10%）    10%')

row = write_sub_objective(row,
    '*1. 个人能力提升指标',
    '围绕知识平台建设的核心技能、AI与新技术应用能力及国际学术素养，系统性提升岗位胜任力',
    '权重 10%')

row = write_col_headers(row)

growth_items = [
    {
        'desc': (
            '知识平台建设专业能力提升\n\n'
            '通过参加行业培训、同类平台研究或自主学习，持续提升在知识平台规划、'
            '内容运营、用户增长等方面的专业能力，年度至少完成2次专业学习或培训'
        ),
        'metrics': (
            '1. 参加专业培训/行业研讨次数（≥2次/年）\n'
            '2. 学习成果转化体现（如运营策略优化、用户增长方法改进等）\n'
            '3. 知识平台建设与运营能力年度对比提升评估'
        ),
        'weight': 4,
    },
    {
        'desc': (
            'AI与新技术应用能力\n\n'
            '结合智识库平台对AI功能的需求，主动学习并掌握AI工具在知识管理、'
            '内容运营、用户体验优化中的应用，提升工作效率与创新产出'
        ),
        'metrics': (
            '1. AI工具在实际工作中的应用场景数（≥2个）\n'
            '2. 新技术应用对工作效率的提升效果\n'
            '3. 在平台建设中的创新方案与技术应用能力表现'
        ),
        'weight': 3,
    },
    {
        'desc': (
            '国际学术视野拓展\n\n'
            '通过阅读国际案例期刊、参与学术交流等方式拓展国际视野，'
            '提升中翻英学术写作与国际投稿能力'
        ),
        'metrics': (
            '1. 国际案例期刊/学术资料阅读量（≥5篇/年）\n'
            '2. 中翻英写作能力提升评估\n'
            '3. 国际投稿流程熟练度与独立操作能力'
        ),
        'weight': 3,
    },
]

row = write_items(row, growth_items)

# ══════════════════════════════════════════════════
# 权重汇总
# ══════════════════════════════════════════════════
row += 1
ws.merge_cells(f'A{row}:C{row}')
ws[f'A{row}'].value = '权重汇总'
ws[f'A{row}'].font = Font(name='微软雅黑', size=11, bold=True)
ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
ws[f'A{row}'].fill = blue_fill
ws[f'A{row}'].border = thin
for c in 'BC':
    ws[f'{c}{row}'].fill = blue_fill
    ws[f'{c}{row}'].border = thin
ws[f'D{row}'].value = 100
ws[f'D{row}'].font = Font(name='微软雅黑', size=11, bold=True, color='C00000')
ws[f'D{row}'].alignment = ctr
ws[f'D{row}'].fill = blue_fill
ws[f'D{row}'].border = thin
ws.row_dimensions[row].height = 30

row += 1
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'].value = (
    '权重分配说明：\n'
    '• 业务指标 70%：*1.智识库平台 65%（平台搭建15% + IT协作15% + 内容管理后台10% + 内容运营10% + 用户测试15%）+ *2.失败案例图谱 5%\n'
    '• 工作态度指标 20%：交付时效与责任心6% + 跨部门协作5% + 主动担当5% + 质量意识4%\n'
    '• 个人能力提升指标 10%：平台建设能力4% + AI与新技术3% + 国际学术视野3%\n'
    '注：项目处于起步阶段，业务指标侧重平台搭建与IT协作两大核心任务'
)
ws[f'A{row}'].font = sm_font
ws[f'A{row}'].alignment = wrap
ws.row_dimensions[row].height = 100

# 打印设置
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = 'landscape'

output_path = r'C:\Users\LXG\fdsmarticles\docs\KPI考核表_智识库项目_v3.xlsx'
wb.save(output_path)
print(f'KPI Excel saved to: {output_path}')
